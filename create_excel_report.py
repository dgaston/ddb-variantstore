#!/usr/bin/env python

import sys
import csv
import utils
import argparse
import getpass

from openpyxl import Workbook
from ddb import configuration
from collections import defaultdict

from variantstore import SampleVariant
from coveragestore import SampleCoverage

from cassandra.cqlengine import connection
from cassandra.auth import PlainTextAuthProvider


def get_target_amplicons(filename):
    amplicons_list = list()
    with open(filename, "r") as bedfile:
        reader = csv.reader(bedfile, dialect='excel-tab')
        for row in reader:
            amplicons_list.append(row[3])

    return amplicons_list


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('-s', '--samples_file', help="Input configuration file for samples")
    parser.add_argument('-c', '--configuration', help="Configuration file for various settings")
    parser.add_argument('-r', '--report', help="Root name for reports (per sample)", default='report')
    parser.add_argument('-a', '--address', help="IP Address for Cassandra connection", default='127.0.0.1')
    parser.add_argument('-u', '--username', help='Cassandra username for login', default=None)

    args = parser.parse_args()

    sys.stdout.write("Parsing configuration data\n")
    config = configuration.configure_runtime(args.configuration)

    sys.stdout.write("Parsing sample data\n")
    libraries = configuration.configure_samples(args.samples_file, config)

    samples = configuration.merge_library_configs_samples(libraries)

    if args.username:
        password = getpass.getpass()
        auth_provider = PlainTextAuthProvider(username=args.username, password=password)
        connection.setup([args.address], "variantstore", auth_provider=auth_provider)
    else:
        connection.setup([args.address], "variantstore")

    thresholds = {'min_saf': 0.01,
                  'max_maf': 0.005,
                  'depth': 200}

    callers = ("mutect", "platypus", "vardict", "scalpel", "freebayes", "pindel")

    sys.stdout.write("Processing samples\n")
    for sample in samples:
        sys.stdout.write("Processing variants for sample {}\n".format(sample))

        wb = Workbook()
        cov_ws = wb.create_sheet(title="Coverage")
        var_ws = wb.create_sheet(title="Variants")
        filter_ws = wb.create_sheet(title="Filtered Variants")

        with open("{}.{}.log".format(sample, args.report), 'w') as logfile:
            logfile.write("Reporting Log for sample {}\n".format(sample))
            logfile.write("---------------------------------------------\n")

        passing_variants = list()
        low_freq_variants = list()
        off_target_variants = list()
        fbayes_pindel_only_variants = list()
        target_amplicon_coverage = defaultdict(lambda: defaultdict(float))

        for library in samples[sample]:
            reportable_amplicons = list()
            off_target_amplicons = defaultdict(int)
            filtered_no_amplicon = list()
            filtered_non_target_amplicon = list()
            filtered_no_requested_caller = list()

            report_panel_path = "/mnt/shared-data/ddb-configs/disease_panels/{}/{}" \
                                "".format(samples[sample][library]['panel'], samples[sample][library]['report'])
            target_amplicons = get_target_amplicons(report_panel_path)

            sys.stdout.write("Processing variants for library {}\n".format(library))
            sys.stdout.write("Processing amplicons for library from file {}\n".format(report_panel_path))

            with open("{}.{}.log".format(sample, args.report), 'a') as logfile:
                logfile.write("Processing variants for library {}\n".format(library))
                logfile.write("Processing amplicons for library from file {}\n".format(report_panel_path))

            variants = SampleVariant.objects.timeout(None).filter(
                SampleVariant.reference_genome == config['genome_version'],
                SampleVariant.sample == samples[sample][library]['sample_name'],
                SampleVariant.run_id == samples[sample][library]['run_id'],
                SampleVariant.library_name == samples[sample][library]['library_name'],
                SampleVariant.max_maf_all <= thresholds['max_maf'],
                ).allow_filtering()

            for amplicon in target_amplicons:
                coverage_data = SampleCoverage.objects.timeout(None).filter(
                    SampleCoverage.sample == samples[sample][library]['sample_name'],
                    SampleCoverage.amplicon == amplicon,
                    SampleCoverage.run_id == samples[sample][library]['run_id'],
                    SampleCoverage.library_name == samples[sample][library]['library_name'],
                    SampleCoverage.program_name == "sambamba"
                )
                ordered_amplicons = coverage_data.order_by('amplicon', 'run_id').limit(coverage_data.count() + 1000)
                for result in ordered_amplicons:
                    reportable_amplicons.append(result)
                    target_amplicon_coverage[amplicon]['num_reads'] = result.num_reads
                    target_amplicon_coverage[amplicon]['mean_coverage'] = result.mean_coverage

            ordered_variants = variants.order_by('library_name', 'chr', 'pos',
                                                 'ref', 'alt').limit(variants.count() + 1000)

            sys.stdout.write("Retrieved {} total variants\n".format(ordered_variants.count()))
            with open("{}.{}.log".format(sample, args.report), 'a') as logfile:
                logfile.write("Retrieved {} total variants\n".format(ordered_variants.count()))

            passing, off_target, low_freq, fbpindel_only, off_target_counts = \
                utils.filter_variants(sample, library, args.report, target_amplicons, callers, ordered_variants)

            sys.stdout.write("Sending {} variants to reporting (filtered {} variants for no amplicon data and {} for "
                             "being in a non-targeted amplicon)\n".format(len(passing_variants),
                                                                          len(filtered_no_amplicon),
                                                                          len(filtered_non_target_amplicon)))

        utils.write_sample_variant_report_no_caller_filter(args.report, sample, passing_variants,
                                                           target_amplicon_coverage, callers)
        wb.save("{}_report.xlsx".format(sample))
